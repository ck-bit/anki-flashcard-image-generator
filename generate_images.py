#!/usr/bin/env python3
"""
Parallel image generation for Korean flashcard images.

Usage:
    pip install fal-client aiohttp
    export FAL_KEY="your-fal-api-key"

    # Generate from session CSV (primary workflow)
    python generate_images.py --csv session_2026-02-23.csv

    # Also still supports xlsx prompt sheets
    python generate_images.py --xlsx korean_anki_image_prompts_v2.xlsx

    # Options
    --images-per-scene 2
    --output flashcard_images
    --concurrency 10
    --limit 5
    --dry-run

Output:
    flashcard_images/{job_id}/
        001_a_거의_{job_id}.png
        001_b_거의_{job_id}.png
        manifest_{job_id}.json
"""

import asyncio
import csv
import json
import re
import argparse
import string
from pathlib import Path
from datetime import datetime

try:
    import fal_client
except ImportError:
    print("Run: pip install fal-client")
    exit(1)


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
MODEL = "fal-ai/flux-pro/v1.1"
IMAGE_SIZE = "square"
MAX_CONCURRENT = 10
IMAGES_PER_SCENE = 2


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def sanitize_filename(text: str, max_len: int = 20) -> str:
    clean = re.sub(r"[^\w가-힣]", "", text)
    return clean[:max_len] if clean else "unknown"


def extract_korean_label(sentence: str) -> str:
    match = re.search(r"[가-힣]+", sentence)
    return match.group(0) if match else "unknown"


# ---------------------------------------------------------------------------
# Prompt loading
# ---------------------------------------------------------------------------
def load_prompts_from_csv(path: str) -> list[tuple[int, str, str]]:
    """Load prompts from session CSV.
    Returns unique [(scene_id, prompt, korean_label), ...].
    Reads 'Prompt A' column. Deduplicates by Scene ID (one prompt per scene)."""
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    seen_scenes = {}
    skipped = []

    for row in rows:
        scene_id = int(row["Scene ID"])
        if scene_id in seen_scenes:
            continue

        prompt = (row.get("Prompt A") or "").strip()
        if not prompt:
            skipped.append(scene_id)
            continue

        sentence = row.get("Full Sentence", "")
        korean_label = extract_korean_label(sentence)
        seen_scenes[scene_id] = (scene_id, prompt, korean_label)

    if skipped:
        print(f"WARNING: {len(skipped)} scenes have no prompt in 'Prompt A': {skipped}")

    return list(seen_scenes.values())


def load_prompts_from_xlsx(path: str) -> list[tuple[int, str, str]]:
    """Load prompts from the image prompt spreadsheet (legacy support).
    Reads column H (Selected Prompt), falls back to column E (Prompt A)."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True)
    ws = wb["Image Prompts"]

    scenes = []
    skipped = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 8:
            continue
        scene_id = row[0]
        if scene_id is None:
            continue

        sentence = (row[1] or "").strip()
        selected = (row[7] or "").strip()
        fallback = (row[4] or "").strip()
        prompt = selected or fallback

        if not prompt:
            skipped.append(scene_id)
            continue

        korean_label = extract_korean_label(sentence)
        scenes.append((int(scene_id), prompt, korean_label))

    wb.close()

    if skipped:
        print(f"WARNING: {len(skipped)} scenes have no prompt: {skipped}")

    return scenes


def load_prompts_from_file(path: str) -> list[tuple[int, str, str]]:
    """Load from text file. Scene IDs assigned sequentially."""
    lines = [
        l.strip() for l in Path(path).read_text().splitlines()
        if l.strip() and not l.startswith("#")
    ]
    return [(i + 1, line, f"scene{i+1}") for i, line in enumerate(lines)]


# ---------------------------------------------------------------------------
# Core
# ---------------------------------------------------------------------------
async def generate_one(
    scene_id: int,
    letter: str,
    prompt: str,
    korean_label: str,
    job_id: str,
    semaphore: asyncio.Semaphore,
) -> dict:
    safe_label = sanitize_filename(korean_label)
    filename = f"{scene_id:03d}_{letter}_{safe_label}_{job_id}.png"
    async with semaphore:
        print(f"  [{filename}] Generating...")
        try:
            result = await asyncio.to_thread(
                fal_client.subscribe,
                MODEL,
                arguments={
                    "prompt": prompt,
                    "image_size": IMAGE_SIZE,
                    "num_images": 1,
                },
            )
            image_url = result["images"][0]["url"]
            print(f"  [{filename}] ✓ Done")
            return {
                "scene_id": scene_id,
                "letter": letter,
                "korean": korean_label,
                "filename": filename,
                "job_id": job_id,
                "prompt": prompt,
                "url": image_url,
                "error": None,
            }
        except Exception as e:
            print(f"  [{filename}] ✗ Failed: {e}")
            return {
                "scene_id": scene_id,
                "letter": letter,
                "korean": korean_label,
                "filename": filename,
                "job_id": job_id,
                "prompt": prompt,
                "url": None,
                "error": str(e),
            }


async def generate_batch(
    scenes: list[tuple[int, str, str]],
    images_per_scene: int,
    job_id: str,
) -> list[dict]:
    semaphore = asyncio.Semaphore(MAX_CONCURRENT)
    letters = string.ascii_lowercase[:images_per_scene]

    tasks = []
    for scene_id, prompt, korean_label in scenes:
        for letter in letters:
            tasks.append(
                generate_one(scene_id, letter, prompt, korean_label, job_id, semaphore)
            )

    return await asyncio.gather(*tasks)


# ---------------------------------------------------------------------------
# Download
# ---------------------------------------------------------------------------
async def download_images(results: list[dict], output_dir: Path):
    import aiohttp

    output_dir.mkdir(parents=True, exist_ok=True)

    async with aiohttp.ClientSession() as session:
        for r in results:
            if not r["url"]:
                continue
            filepath = output_dir / r["filename"]
            async with session.get(r["url"]) as resp:
                with open(filepath, "wb") as f:
                    f.write(await resp.read())
            print(f"  Saved {filepath}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
async def main():
    global MAX_CONCURRENT, MODEL

    parser = argparse.ArgumentParser(description="Generate flashcard images")
    parser.add_argument("--csv", help="Session CSV file (primary input)")
    parser.add_argument("--xlsx", help="Image prompts spreadsheet (legacy)")
    parser.add_argument("-f", "--file", help="Prompts text file (one per line)")
    parser.add_argument("-o", "--output", default="flashcard_images", help="Output base directory")
    parser.add_argument("-n", "--concurrency", type=int, default=MAX_CONCURRENT)
    parser.add_argument("--images-per-scene", type=int, default=IMAGES_PER_SCENE)
    parser.add_argument("--model", default=MODEL, help=f"FAL model (default: {MODEL})")
    parser.add_argument("--no-download", action="store_true", help="Skip downloading")
    parser.add_argument("--dry-run", action="store_true", help="Preview without generating")
    parser.add_argument("--limit", type=int, default=None, help="Only first N scenes")
    args = parser.parse_args()

    MAX_CONCURRENT = args.concurrency
    MODEL = args.model

    job_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = Path(args.output) / job_id

    # Load prompts
    if args.csv:
        scenes = load_prompts_from_csv(args.csv)
    elif args.xlsx:
        scenes = load_prompts_from_xlsx(args.xlsx)
    elif args.file:
        scenes = load_prompts_from_file(args.file)
    else:
        print("Provide --csv <session.csv>, --xlsx <spreadsheet>, or -f <prompts.txt>")
        return

    if args.limit:
        scenes = scenes[:args.limit]

    total_images = len(scenes) * args.images_per_scene

    print(f"\n🎴 {len(scenes)} scenes × {args.images_per_scene} = {total_images} images")
    print(f"   Model:   {MODEL}")
    print(f"   Job ID:  {job_id}")
    print(f"   Output:  {output_dir}\n")

    if args.dry_run:
        letters = string.ascii_lowercase[:args.images_per_scene]
        for scene_id, prompt, korean_label in scenes:
            safe = sanitize_filename(korean_label)
            fnames = ", ".join(f"{scene_id:03d}_{l}_{safe}_{job_id}.png" for l in letters)
            print(f"  Scene {scene_id:03d} ({korean_label})")
            print(f"    files: {fnames}")
            print(f"    prompt: {prompt[:90]}...")
            print()
        print(f"Dry run complete. {total_images} images would be generated.")
        return

    results = await generate_batch(scenes, args.images_per_scene, job_id)

    succeeded = [r for r in results if r["url"]]
    failed = [r for r in results if r["error"]]
    print(f"\n✓ {len(succeeded)} succeeded, ✗ {len(failed)} failed")
    print(f"  Job ID: {job_id}\n")

    output_dir.mkdir(parents=True, exist_ok=True)
    manifest_path = output_dir / f"manifest_{job_id}.json"
    manifest_path.write_text(json.dumps(results, indent=2, ensure_ascii=False))
    print(f"Manifest saved to {manifest_path}")

    if not args.no_download and succeeded:
        print("\nDownloading images...")
        await download_images(results, output_dir)

    if failed:
        print(f"\nFailed scenes:")
        for r in failed:
            print(f"  {r['filename']}: {r['error']}")

    print(f"\nDone! 🎉  Job ID: {job_id}")


if __name__ == "__main__":
    asyncio.run(main())